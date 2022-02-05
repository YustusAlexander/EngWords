import random
import openpyxl
from openpyxl.styles import Font

path_to_file = 'My words ENG.xlsx'
try:
    wb = openpyxl.load_workbook(path_to_file)
    ws = wb.active

except:
    print("Документ открыт")


def available_row():
    return [i for i in range(ws.min_row, ws.max_row + 1) if ws['A' + str(i)].value is not None]

def available_fav_row():
    return [i for i in range(ws.min_row, ws.max_row + 1) if ws['A' + str(i)].value is not None
            and ws.cell(row=i, column=1).font == Font(bold=True)]

def output_row(range_row):
    count = 0
    for i in available_row():
        if count <= range_row:
            count += 1
            return "\n".join(ws.cell(row=ws.min_row + i, column=1).value.strip()
                             + ' - ' + ws.cell(row=ws.min_row + i, column=2).value.strip())

########################################################################################

def last(num):
    # return output_row(val_row)
    return "\n".join([ws.cell(row=ws.min_row+i, column=1).value.strip()
                      + ' - ' + ws.cell(row=ws.min_row+i, column=2).value.strip() for i in available_row() if i < num])


def rand(num):
    val_row = random.sample(available_row(), num)
    # return output_row(val_row)
    return "\n".join([ws.cell(row=i, column=1).value.strip()
                      + ' - ' + ws.cell(row=i, column=2).value.strip() for i in val_row])


def all_list():
    return "\n".join([ws.cell(row=i, column=1).value.strip()
                      + ' - ' + ws.cell(row=i, column=2).value.strip() for i in available_row()])


def list_num():
    return "\n".join([str(i) + '. ' + ws.cell(row=i, column=1).value.strip()
                      + ' - ' + ws.cell(row=i, column=2).value.strip() for i in available_row()])


def favour():
    return "\n".join([ws.cell(row=i, column=1).value.strip()
                      + ' - ' + ws.cell(row=i, column=2).value.strip() for i in available_row()
                      if ws.cell(row=i, column=1).font == Font(bold=True)])

def favour_num():
    return "\n".join([str(i) + '. ' + ws.cell(row=i, column=1).value.strip()
                      + ' - ' + ws.cell(row=i, column=2).value.strip() for i in available_fav_row()])



###############################################

def match():
    return "\n".join([ws.cell(row=ws.min_row + i, column=1).value.strip()
                      + ' - ' + ws.cell(row=ws.min_row + i, column=2).value.strip() for i in range(20)])
